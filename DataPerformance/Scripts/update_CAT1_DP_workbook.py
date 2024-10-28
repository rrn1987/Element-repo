import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, styles
from datetime import datetime

SHEET_NAME_START = 'LTE_37.901_'
masterfile_path = os.getcwd() + '\\MasterFile\\'
results_path = os.getcwd() + '\\Results\\'
CAT1_file_path = os.path.join(masterfile_path, "Element_(CAT-1)_LTE_Data Performance.xlsx")
Script_Version = IMEI = ROM = None
Test_Platform = Test_Lab = None


def update_sheet(project_workbook, sheet_name, test_id, subtest, header, final_verdict, observation):
    try:
        worksheet = project_workbook[sheet_name]

        header_row = worksheet[header]

        # Find the column indices based on column names
        test_id_column_index = subtest_column_index = result_column_index = None
        KPI_column_index = date_column_index = IMEI_column_index = ROM_column_index = sv_column_index = None

        for cell in header_row:
            if cell.value == 'TMO Test #':
                test_id_column_index = cell.column
            elif cell.value == 'Subtest':
                subtest_column_index = cell.column
            elif cell.value == 'Result':
                result_column_index = cell.column
            elif cell.value == 'Throughput':
                KPI_column_index = cell.column
            elif cell.value == 'Test Date':
                date_column_index = cell.column
            elif cell.value == 'Test Platform':
                test_platform_column_index = cell.column
            elif cell.value == 'Test Lab':
                test_lab_column_index = cell.column
            elif cell.value == 'IMEI':
                IMEI_column_index = cell.column
            elif cell.value == 'ROM':
                ROM_column_index = cell.column
            elif cell.value == 'Script Version':
                sv_column_index = cell.column

        # print(str(test_id_column_index) + '-' + str(subtest_column_index) + '-' + str(result_column_index) + '-' +

        row_num = 1
        # Iterate through each row in the worksheet
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_num += 1
            test_id_cell = row[test_id_column_index - 1]
            subtest_cell = row[subtest_column_index - 1]
            result_cell = worksheet.cell(row=row_num, column=result_column_index)
            KPI_cell = worksheet.cell(row=row_num, column=KPI_column_index)
            date_cell = worksheet.cell(row=row_num, column=date_column_index)
            test_platform_cell = worksheet.cell(row=row_num, column=test_platform_column_index)
            test_lab_cell = worksheet.cell(row=row_num, column=test_lab_column_index)
            IMEI_cell = worksheet.cell(row=row_num, column=IMEI_column_index)
            ROM_cell = worksheet.cell(row=row_num, column=ROM_column_index)
            sv_cell = worksheet.cell(row=row_num, column=sv_column_index)

            test_id_cell = str(test_id_cell).replace('.', '')

            # Check if the test_id and subtest values match
            if str(test_id_cell).strip() == str(test_id).strip() and (
                    str(subtest_cell).strip() == str(subtest).strip() or subtest is None):
                result_cell.value = final_verdict
                KPI_cell.value = observation
                date_cell.value = datetime.now().strftime('%m/%d/%Y')
                test_platform_cell.value = Test_Platform
                test_lab_cell.value = Test_Lab
                IMEI_cell.value = IMEI
                sv_cell.value = Script_Version
                ROM_cell.value = ROM
                if final_verdict == 'Pass' or final_verdict == 'Passed':
                    result_cell.fill = styles.PatternFill("solid", start_color="5cb800")
                elif final_verdict == 'Fail':
                    result_cell.fill = styles.PatternFill("solid", start_color="00FF0000")
                break

        # Save the changes to the Excel file
        project_workbook.save(os.path.join(results_path, project_name + "_CAT1_DP.xlsx"))
    except Exception as e:
        print("Error occurred while saving the workbook:", str(e))
    return project_workbook


def get_results(log_path, workbook):
    """
       extract all the testcases in the folder
       """
    project_workbook = workbook
    sheet_name = test_id = subtest = final_verdict = observation = band_split = None
    count = pass_count = 0

    if os.path.exists(log_path):
        print("Updating Results for " + log_path)
        for file in os.listdir(log_path):
            if file.startswith('TMO_'):
                # print(str(count) + "." + file)
                SummaryReport = 'SummaryReport.xml'
                for root, dirs, files in os.walk(log_path + "\\" + file):
                    if SummaryReport in files:
                        count += 1
                        header = 1
                        tree = ET.parse(log_path + "\\" + file + "\\" + SummaryReport)
                        for element in tree.iter(tag='header'):
                            subtest = None
                            testcase_number = element.find('testcasenumber').text
                            observation = element.find('observation').text
                            final_verdict = element.find('finalverdict').text
                            if final_verdict is not None and (final_verdict == 'Pass' or final_verdict == 'Passed'):
                                final_verdict = 'Passed'
                                pass_count += 1
                            # print(str(testcase_number) + '- ' + str(final_verdict) + '- ' + str(observation))
                            if observation is not None:
                                observation = observation.replace('\n', ' ')
                            separate = testcase_number.split('-')
                            # print(len(separate))
                            test_id = separate[2]
                            sheet_name = 'Cat-1'
                            # print(test_id)
                            if len(separate) == 4:
                                subtest = float(separate[3].replace('T', ''))
                                if subtest.is_integer():
                                    subtest = int(subtest)
                                # print(subtest)

                        print(str(count) + '. Tab-' + str(sheet_name) + '\t Test#-' + str(test_id) + '\t Subtest#-' +
                              str(subtest) + '\t Result-' + str(final_verdict) + '\t KPI-' + str(observation))
                        test_id = test_id.replace('.', '')

                        if final_verdict != 'Aborted' and observation != 'Insufficient number of CMW(s) available ' \
                                                                         'in Test System. Required CMW(s):- 3, ' \
                                                                         'Available CMW(s):-2':
                            project_workbook = update_sheet(project_workbook, sheet_name, test_id, subtest, header,
                                                            final_verdict,
                                                            observation)

                    else:
                        pass
                        # print('SummaryReport Not Found')
    print('Total Pass in ' + str(log_path) + ' - ' + str(pass_count))
    project_workbook.save(r'D:\Projects\\' + str(project_name) + '_CAT1_DP.xlsx')


if __name__ == '__main__':
    # log_path = r'D:\logs\DP_LTE_2023-04-25T113125'
    project_name = input("Enter Project Name: ")
    if os.path.exists(project_name + '_CAT1_DP.xlsx'):
        os.remove(project_name + '_CAT1_DP.xlsx')
    master_workbook = load_workbook(CAT1_file_path)
    master_workbook.save(os.path.join(results_path, project_name + "_CAT1_DP.xlsx"))
    IMEI = input("Enter DUT IMEI: ")
    Test_Platform = 'R&S PQA'
    Test_Lab = 'ELEMENT'
    Script_Version = 'PQA7.80'
    ROM = input("Enter ROM: ")
    log_path = input("Enter Result path: ")
    if ';' in log_path:
        log_path = log_path.split(';')
        for each_log_path in log_path:
            get_results(each_log_path, master_workbook)
    else:
        get_results(log_path, master_workbook)

    print('Updated Results sheet Saved to- ' + 'D:\Projects\\' + str(project_name) + '_CAT1_DP.xlsx')


#   D:\logs\T-Mobile US LTE Data Performance Cat-1_2023-06-19T144620;D:\logs\T-Mobile US LTE Data Performance Cat-1_2023-06-19T125056