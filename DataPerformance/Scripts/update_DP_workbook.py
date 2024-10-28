import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, styles
from datetime import datetime

SHEET_NAME_START = 'LTE_37.901_'
masterfile_path = os.getcwd() + '\\MasterFile\\'
results_path = os.getcwd() + '\\Results\\'
file_path = os.path.join(masterfile_path, "Element_LTE_DataPerformance.xlsx")
Tester = IMEI = None


def update_sheet(project_workbook, sheet_name, test_id, subtest, header, final_verdict, observation):
    try:
        worksheet = project_workbook[sheet_name]

        header_row = worksheet[header]

        # Find the column indices based on column names
        test_id_column_index = subtest_column_index = result_column_index = KPI_column_index = date_column_index = \
            IMEI_column_index = tester_column_index = None

        yellow = "00FFFF00"
        red = '00FF0000'
        green = '5cb800'

        for cell in header_row:
            if cell.value == 'TMO Test #':
                test_id_column_index = cell.column
            elif cell.value == 'Subtest':
                subtest_column_index = cell.column
            elif cell.value == 'Result Verdict':
                result_column_index = cell.column
            elif cell.value == 'Comments\nTHROUGHPUT VALUE':
                KPI_column_index = cell.column
            elif cell.value == 'Date':
                date_column_index = cell.column
            elif cell.value == 'DUT IMEI':
                IMEI_column_index = cell.column
            elif cell.value == 'Tester':
                tester_column_index = cell.column

        # print(str(test_id_column_index) + '-' + str(subtest_column_index) + '-' + str(result_column_index) + '-' +
        # str(KPI_column_index))
        row_num = 1
        # Iterate through each row in the worksheet
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_num += 1

            test_id_cell = row[test_id_column_index - 1]
            if subtest_column_index is not None:
                subtest_cell = row[subtest_column_index - 1]
            else:
                subtest_cell = None
            result_cell = worksheet.cell(row=row_num, column=result_column_index)
            KPI_cell = worksheet.cell(row=row_num, column=KPI_column_index)
            date_cell = worksheet.cell(row=row_num, column=date_column_index)
            IMEI_cell = worksheet.cell(row=row_num, column=IMEI_column_index)
            tester_cell = worksheet.cell(row=row_num, column=tester_column_index)

            test_id_cell = str(test_id_cell).replace('.', '')

            # Check if the test_id and subtest values match
            if str(test_id_cell).strip() == str(test_id).strip() and (
                    str(subtest_cell).strip() == str(subtest).strip() or subtest is None):
                result_cell.value = final_verdict
                KPI_cell.value = observation
                date_cell.value = datetime.now().strftime('%m/%d/%Y')
                IMEI_cell.value = IMEI
                tester_cell.value = Tester
                if final_verdict == 'Pass' or final_verdict == 'Passed':
                    result_cell.fill = styles.PatternFill(fill_type="solid", start_color=green)
                elif final_verdict == 'Fail':
                    result_cell.fill = styles.PatternFill(fill_type="solid", start_color=red)
                elif final_verdict == 'Inconclusive':
                    result_cell.fill = styles.PatternFill(fill_type="solid", start_color=yellow)
                break

        # Save the changes to the Excel file
        project_workbook.save(os.path.join(results_path, project_name + "_DP.xlsx"))
        # print("Changes saved successfully.")
    except Exception as e:
        print("Error occurred while saving the workbook:", str(e))
    return project_workbook


def get_results(log_path, workbook):
    """
       extract all the testcases in the folder
       """
    project_workbook = workbook
    sheet_name = test_id = subtest = final_verdict = observation = band_split = None
    count = pass_count = 0
    CA = 'CA'

    if os.path.exists(log_path):
        print("Updating Results for " + log_path)
        for file in os.listdir(log_path):
            if file.startswith('TMO_'):
                # print(str(count) + "." + file)
                SummaryReport = 'SummaryReport.xml'
                for root, dirs, files in os.walk(log_path + "\\" + file):
                    if SummaryReport in files:
                        header = 1
                        try:
                            tree = ET.parse(log_path + "\\" + file + "\\" + SummaryReport)
                            for element in tree.iter(tag='header'):
                                subtest = None
                                testcase_number = element.find('testcasenumber').text
                                observation = element.find('observation').text
                                final_verdict = element.find('finalverdict').text
                                band = element.find('band').text
                                if final_verdict is not None and (final_verdict == 'Pass' or final_verdict == 'Passed'):
                                    final_verdict = 'Passed'
                                    pass_count += 1
                                if band is not None:
                                    band_split = band.split('_')
                                # print(str(testcase_number) + '- ' + str(final_verdict) + '- ' + str(band))
                                if observation is not None:
                                    observation = observation.replace('\n', ' ')
                                separate = testcase_number.split('-')
                                # print(separate)

                                if testcase_number is not None and final_verdict != 'Error' and final_verdict != \
                                        'Not Applicable' and final_verdict != 'Not Initialized' and \
                                        final_verdict != 'Aborted':
                                    count += 1
                                    if band_split[0] == CA and separate[1] != 'Band_41' and '48' not in band:
                                        sheet_name = CA
                                        header = 3
                                        test_id = separate[1]
                                        if len(separate) == 3:
                                            subtest = float(separate[2].replace('T', ''))
                                            if subtest.is_integer():
                                                subtest = int(subtest)
                                    else:
                                        test_id = separate[2]
                                        if test_id[0] == '3' and len(separate) != 4:
                                            sheet_name = '4x2 MIMO'
                                            test_id_split = str(test_id).split('.')
                                            test_id = test_id_split[0] + test_id_split[1]
                                            if len(test_id_split) == 2:
                                                subtest = None
                                            elif len(test_id_split) == 3:
                                                subtest = int(test_id_split[2])
                                        elif test_id[0] == '6' and len(separate) != 4:
                                            sheet_name = '4x4 MIMO'
                                            test_id_split = str(test_id).split('.')
                                            if len(test_id_split) == 2:
                                                test_id = test_id_split[0] + test_id_split[1]
                                                subtest = None
                                            elif len(test_id_split) == 3:
                                                test_id = test_id_split[0] + test_id_split[1]
                                                subtest = int(test_id_split[2])
                                            elif len(test_id_split) == 4:
                                                test_id = test_id_split[0] + test_id_split[1] + test_id_split[2]
                                                subtest = int(test_id_split[3])
                                        elif str(test_id).startswith('480') and len(separate) != 4:
                                            sheet_name = 'Band 48'
                                        else:
                                            sheet_name = separate[1].replace('_', ' ')

                                            # print(test_id)
                                        if len(separate) == 4:
                                            if test_id[0] == '3':
                                                sheet_name = '4x2 MIMO'
                                            elif test_id[0] == '6':
                                                sheet_name = '4x4 MIMO'
                                            elif str(test_id).startswith('480'):
                                                sheet_name = 'Band 48'
                                            subtest = float(separate[3].replace('T', ''))
                                            if subtest.is_integer():
                                                subtest = int(subtest)
                                            # print(subtest)
                                        if sheet_name == 'Band 02':
                                            header = 3
                                            sheet_name = SHEET_NAME_START + 'Band 02'
                                        elif sheet_name == 'Band 04':
                                            sheet_name = SHEET_NAME_START + 'Band 04'
                                        elif sheet_name == 'Band 05':
                                            sheet_name = SHEET_NAME_START + 'Band 05'
                                        elif sheet_name == 'Band 12':
                                            header = 2
                                            sheet_name = SHEET_NAME_START + 'Band 12'
                                        elif sheet_name == 'Band 66':
                                            header = 2
                                            sheet_name = SHEET_NAME_START + 'Band 66'
                                        elif sheet_name == 'Band 71':
                                            sheet_name = 'Band 71'
                                        elif sheet_name == 'Band 41':
                                            sheet_name = 'Band 41'

                                        print(str(count) + '. Tab-' + str(sheet_name) + '\t Test#-' + str(
                                            test_id) + '\t Subtest#-' +
                                              str(subtest) + '\t Result-' + str(final_verdict) + '\t KPI-' + str(
                                                    observation))
                                        test_id = test_id.replace('.', '')

                                        if observation != 'Insufficient number of CMW(s) available ' \
                                                          'in Test System. Required CMW(s):- 3, ' \
                                                          'Available CMW(s):-2':
                                            project_workbook = update_sheet(project_workbook, sheet_name, test_id,
                                                                            subtest, header,
                                                                            final_verdict,
                                                                            observation)
                        except ET.ParseError as e:
                            print("XML parsing error:", str(e))

                    else:
                        pass
                        # print('SummaryReport Not Found')
    print('Total Pass in ' + str(log_path) + ' - ' + str(pass_count))
    project_workbook.save(r'D:\Projects\\' + project_name + '_DP.xlsx')
    return pass_count


if __name__ == '__main__':
    total_pass_count = 0
    project_name = input("Enter Project Name: ")
    if os.path.exists(project_name + '_DP.xlsx'):
        os.remove(project_name + '_DP.xlsx')
    master_workbook = load_workbook(file_path)
    master_workbook.save(os.path.join(results_path, project_name + "_DP.xlsx"))
    IMEI = input("Enter DUT IMEI: ")
    Tester = 'FK'
    log_path = input("Enter Result path: ")
    if ';' in log_path:
        log_path = log_path.split(';')
        for each_log_path in log_path:
            total_pass_count += get_results(each_log_path, master_workbook)
    else:
        total_pass_count = get_results(log_path, master_workbook)
    print('Overall Pass: ' + str(total_pass_count))
    print('Updated Results sheet Saved to- ' + 'D:\Projects\\' + str(project_name) + '_DP.xlsx')
