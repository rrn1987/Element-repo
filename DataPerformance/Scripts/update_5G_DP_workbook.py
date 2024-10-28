import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, styles
from datetime import datetime
import bs4 as bs

SHEET_NAME_START = 'LTE_37.901_'
masterfile_path = r'D:\Users\nairr\PycharmProjects\pythonProject\DataPerformance' + '\\MasterFile\\'
results_path = r'D:\Users\nairr\PycharmProjects\pythonProject\DataPerformance' + '\\Results\\'
FiveG_file_path = os.path.join(masterfile_path, "5G_Data_Performance.xlsx")
Script_Version = IMEI = ROM = None
Test_Platform = Test_Lab = None


def update_sheet(project_workbook, sheet_name, test_id, subtest, header, final_verdict, observation, throughput_limit):
    try:
        worksheet = project_workbook[sheet_name]

        header_row = worksheet[header]

        # Find the column indices based on column names
        test_id_column_index = subtest_column_index = result_column_index = None
        KPI_column_index = date_column_index = IMEI_column_index = ROM_column_index = sv_column_index = None
        test_platform_column_index = test_lab_column_index = throughput_limit_index = None

        yellow = "00FFFF00"
        red = '00FF0000'
        green = '5cb800'

        for cell in header_row:
            if cell.value == 'TMO Test #':
                test_id_column_index = cell.column
            elif cell.value == 'Subset':
                subtest_column_index = cell.column
            elif cell.value == 'Result':
                result_column_index = cell.column
            elif cell.value == 'Throughput':
                KPI_column_index = cell.column
            elif cell.value == 'Throughput Limit':
                throughput_limit_index = cell.column
            elif cell.value == 'Test Date':
                date_column_index = cell.column
            elif cell.value == 'Test Platform':
                test_platform_column_index = cell.column
            elif cell.value == 'Test Lab':
                test_lab_column_index = cell.column
            elif cell.value == 'IMEI':
                IMEI_column_index = cell.column
            elif cell.value == 'ROM':
                ROM_column_index = cell.column
            elif cell.value == 'Script Version':
                sv_column_index = cell.column

        # print(str(test_id_column_index) + '-' + str(subtest_column_index) + '-' + str(result_column_index))

        row_num = 1
        # Iterate through each row in the worksheet
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_num += 1
            test_id_cell = row[test_id_column_index - 1]
            subtest_cell = row[subtest_column_index - 1]
            result_cell = worksheet.cell(row=row_num, column=result_column_index)
            KPI_cell = worksheet.cell(row=row_num, column=KPI_column_index)
            throughput_limit_cell = worksheet.cell(row=row_num, column=throughput_limit_index)
            date_cell = worksheet.cell(row=row_num, column=date_column_index)
            test_platform_cell = worksheet.cell(row=row_num, column=test_platform_column_index)
            test_lab_cell = worksheet.cell(row=row_num, column=test_lab_column_index)
            IMEI_cell = worksheet.cell(row=row_num, column=IMEI_column_index)
            ROM_cell = worksheet.cell(row=row_num, column=ROM_column_index)
            sv_cell = worksheet.cell(row=row_num, column=sv_column_index)

            test_id_cell = str(test_id_cell).replace('.', '')

            # Check if the test_id and subtest values match
            if str(test_id_cell).strip() == str(test_id).strip() and (
                    str(subtest_cell).strip() == str(subtest).strip() or subtest is None):
                result_cell.value = final_verdict
                KPI_cell.value = observation
                throughput_limit_cell.value = throughput_limit
                date_cell.value = datetime.now().strftime('%m/%d/%Y')
                test_platform_cell.value = Test_Platform
                test_lab_cell.value = Test_Lab
                IMEI_cell.value = IMEI
                sv_cell.value = Script_Version
                ROM_cell.value = ROM
                if final_verdict == 'Pass' or final_verdict == 'Passed':
                    result_cell.fill = styles.PatternFill(fill_type="solid", start_color=green)
                elif final_verdict == 'Fail':
                    result_cell.fill = styles.PatternFill(fill_type="solid", start_color=red)
                elif final_verdict == 'Inconclusive':
                    result_cell.fill = styles.PatternFill(fill_type="solid", start_color=yellow)
                break

        # Save the changes to the Excel file
        project_workbook.save(os.path.join(results_path, project_name + "_5G_DP.xlsx"))
    except Exception as e:
        print("Error occurred while saving the workbook:", str(e))
    return project_workbook


def get_KPI(url):
    key_Appl_Tput = key_Appl_Tput_Limit = Appl_Tput = Appl_Tput_Limit = None
    kpi = throughput_limit = None
    file = bs.BeautifulSoup(open(url).read(), "lxml")
    find_table = file.find_all('div', id='table_11_section_1')
    for i in find_table:
        table_head = i.find_all('th')
        head = [j.text for j in table_head]
        table_data = i.find_all('td')
        data = [j.text for j in table_data]
        result_dict = dict(map(lambda a, b: (a, b), head, data))
        # print(result_dict)
        out_dict = dict(list(result_dict.items())[0: 6])  # Get the 1st two Key Value for DL and UL KPIs
        # print(out_dict)
        for key in out_dict:
            kpi = str(out_dict[key])
            if kpi != '-':
                if key == 'DL Appl.Tput' or key == 'UL Appl.Tput':
                    key_Appl_Tput = key
                    Appl_Tput = kpi
                elif key == 'DL Appl.Tput Limit' or key == 'UL Appl.Tput Limit':
                    key_Appl_Tput_Limit = key
                    Appl_Tput_Limit = kpi
        return key_Appl_Tput, key_Appl_Tput_Limit, Appl_Tput, Appl_Tput_Limit


def get_results(log_path, workbook):
    """
       extract all the testcases in the folder
       """
    project_workbook = workbook
    sheet_name = test_id = subtest = final_verdict = observation = band_split = None
    count = pass_count = 0

    if os.path.exists(log_path):
        print("Updating Results for " + log_path)
        for file in os.listdir(log_path):
            if file.startswith('Nr_KN511_TMO'):
                # print(str(count) + "." + file)
                OnlineReport = 'OnlineReport.htm'
                SummaryReport = 'SummaryReport.xml'
                for root, dirs, files in os.walk(log_path + "\\" + file):
                    if SummaryReport in files and OnlineReport in files:
                        url = str(log_path + "\\" + file + "\\" + OnlineReport)
                        key_Appl_Tput, key_Appl_Tput_Limit, Appl_Tput, Appl_Tput_Limit = get_KPI(url)
                        header = 1
                        tree = ET.parse(log_path + "\\" + file + "\\" + SummaryReport)
                        for element in tree.iter(tag='header'):
                            subtest = None
                            testcase_number = element.find('testcasenumber').text
                            final_verdict = element.find('finalverdict').text
                            if testcase_number is not None and final_verdict != 'Error' and final_verdict != \
                                    'Not Applicable' and final_verdict != 'Not Initialized' and \
                                    final_verdict != 'Aborted' and Appl_Tput is not None:
                                count += 1
                                if final_verdict is not None and (final_verdict == 'Pass' or final_verdict == 'Passed'):
                                    final_verdict = 'Passed'
                                    pass_count += 1
                                # print(str(testcase_number) + '- ' + str(final_verdict) + '- ' + str(observation))
                                separate = testcase_number.split('-')
                                # print(len(separate))
                                test_id = separate[7]
                                sheet_name = 'All TCs result(V23.2)'
                                # print(test_id)
                                if len(separate) == 9:
                                    subtest = float(separate[8].replace('T', ''))
                                    if subtest.is_integer():
                                        subtest = int(subtest)
                                    # print(subtest)

                            print(
                                str(count) + '. Tab-' + str(sheet_name) + '\t Test#-' + str(test_id) + '\t Subtest#-' +
                                str(subtest) + '\t Result-' + str(final_verdict) + '\t' + str(
                                    key_Appl_Tput) + '-' + str(Appl_Tput) + '\t' + str(key_Appl_Tput_Limit) + '-' + str(Appl_Tput_Limit))
                            test_id = test_id.replace('.', '')

                            project_workbook = update_sheet(project_workbook, sheet_name, test_id, subtest, header,
                                                            final_verdict,
                                                            Appl_Tput, Appl_Tput_Limit)

                    else:
                        pass
                        # print('SummaryReport Not Found')
    print('Total Pass in ' + str(log_path) + ' - ' + str(pass_count))
    project_workbook.save(r'D:\Projects\\' + str(project_name) + '_5G_DP.xlsx')


if __name__ == '__main__':
    log_path = r'D:\Projects\DP Automation\DP_log\TMO US FR1 SA Data Throughput_2023-07-13T131800'
    project_name = 'test'  # input("Enter Project Name: ")
    if os.path.exists(project_name + '_5G_DP.xlsx'):
        os.remove(project_name + '_5G_DP.xlsx')
    master_workbook = load_workbook(FiveG_file_path)
    master_workbook.save(os.path.join(results_path, project_name + "_5G_DP.xlsx"))
    IMEI = '1414213'
    ROM = '3'
    # IMEI = input("Enter DUT IMEI: ")
    Test_Platform = 'R&S PQA'
    Test_Lab = 'ELEMENT'
    Script_Version = 'PQA7.80'
    # ROM = input("Enter ROM: ")
    # log_path = input("Enter Result path: ")
    if ';' in log_path:
        log_path = log_path.split(';')
        for each_log_path in log_path:
            get_results(each_log_path, master_workbook)
    else:
        get_results(log_path, master_workbook)

    print('Updated Results sheet Saved to- ' + 'D:\Projects\\' + str(project_name) + '_5G_DP.xlsx')
